from random import random
from tkinter import *
from tkinter import ttk
from tkinter import filedialog
import tkinter.messagebox as mb
import openpyxl


def view():
    text_staff.delete('1.0', END)
    x = ''
    for i in range(0, ws.max_row):
        for col in ws.iter_cols(1, ws.max_column):
            if col[i].value != None:
                x += f'{col[i].value} '
        if x:
            text_staff.insert(END, f'{i+1}. {x}\n')
        x = ''
    global rows_count
    rows_count = len([row for row in ws if not all(
        [cell.value is None for cell in row])])


def open_file():
    filetypes = (("Excel файл", "*.xlsx"), ("Любой", "*"))
    filepath = filedialog.askopenfilename(filetypes=filetypes)
    global filename
    filename = filepath[filepath.rfind('/') + 1:]
    global wb
    wb = openpyxl.load_workbook(filepath)
    global ws
    ws = wb.active
    global rows_count
    rows_count = len([row for row in ws if not all(
        [cell.value is None for cell in row])])
    view()
    combobox.place(x=705, y=200)
    combobox.bind("<<ComboboxSelected>>", selected)
    open_btn.place_forget()


def selected(event):
    lastname.place_forget()
    firstname.place_forget()
    age.place_forget()
    num_employee.place_forget()
    fathername.place_forget()
    one_let.place_forget()
    lastname.delete(0, END)
    firstname.delete(0, END)
    fathername.delete(0, END)
    age.delete(0, END)
    num_employee.delete(0, END)
    add_label.place_forget()
    lastname_label.place_forget()
    firstname_label.place_forget()
    fathername_label.place_forget()
    age_label.place_forget()
    del_label.place_forget()
    num_employees_label.place_forget()
    update_label.place_forget()
    search_lastname_label.place_forget()
    search_one_let_label.place_forget()
    one_let_label.place_forget()
    search_age_label.place_forget()
    update_helper.place_forget()
    save_btn.place_forget()
    add_btn.place_forget()
    del_btn.place_forget()
    update_btn.place_forget()
    search_btn.place_forget()
    search_one_let_btn.place_forget()
    search_age_btn.place_forget()
    view_btn.place_forget()

    if combobox.get() == 'Добавить':
        add_label.place(x=680, y=0)
        lastname_label.place(x=690, y=30)
        lastname.place(width=100, height=20, x=750, y=30)
        firstname_label.place(x=690, y=55)
        firstname.place(width=100, height=20, x=750, y=55)
        fathername_label.place(x=690, y=80)
        fathername.place(width=100, height=20, x=750, y=80)
        age_label.place(x=690, y=105)
        age.place(width=100, height=20, x=750, y=105)
        add_btn.place(width=100, height=20, x=725, y=150)
    if combobox.get() == 'Редактировать':
        update_label.place(x=680, y=0)
        num_employees_label.place(x=690, y=30)
        num_employee.place(width=50, height=20, x=810, y=30)
        lastname_label.place(x=690, y=60)
        lastname.place(width=100, height=20, x=750, y=60)
        firstname_label.place(x=690, y=85)
        firstname.place(width=100, height=20, x=750, y=85)
        fathername_label.place(x=690, y=110)
        fathername.place(width=100, height=20, x=750, y=110)
        age_label.place(x=690, y=135)
        age.place(width=100, height=20, x=750, y=135)
        update_btn.place(width=100, height=20, x=725, y=170)
        update_helper.place(x=675, y=230)
    if combobox.get() == 'Удалить':
        del_label.place(x=680, y=0)
        num_employees_label.place(x=690, y=30)
        num_employee.place(width=50, height=20, x=820, y=30)
        del_btn.place(width=100, height=20, x=725, y=150)
    if combobox.get() == 'Найти сотрудника':
        search_lastname_label.place(x=680, y=0)
        lastname_label.place(x=690, y=30)
        lastname.place(width=100, height=20, x=750, y=30)
        search_btn.place(width=100, height=20, x=725, y=150)
    if combobox.get() == 'Найти по первой букве':
        search_one_let_label.place(x=690, y=0)
        one_let_label.place(x=690, y=30)
        one_let.place(width=100, height=20, x=750, y=30)
        search_one_let_btn.place(width=100, height=20, x=725, y=150)
    if combobox.get() == 'Найти по возрасту':
        search_age_label.place(x=690, y=0)
        age_label.place(x=690, y=30)
        age.place(width=100, height=20, x=750, y=30)
        search_age_btn.place(width=100, height=20, x=725, y=150)


def add():
    view()
    if lastname.get() and firstname.get() and fathername.get() and age.get():
        a = ws.cell(row=rows_count + 1, column=1)
        a.value = lastname.get()
        b = ws.cell(row=rows_count + 1, column=2)
        b.value = firstname.get()
        c = ws.cell(row=rows_count + 1, column=3)
        c.value = fathername.get()
        d = ws.cell(row=rows_count + 1, column=4)
        d.value = age.get()
        wb.save(filename)
        mb.showinfo('Информация', 'Сотрудник успешно добавлен!')
        lastname.delete(0, END)
        firstname.delete(0, END)
        fathername.delete(0, END)
        age.delete(0, END)
        view()
    else:
        mb.showerror('ОШИБКА', 'Все поля должны быть заполнены!')


def choice_update():
    view()
    if num_employee.get():
        global choice
        choice = int(num_employee.get())
        if choice <= int(rows_count) and choice > 0:
            a = ws.cell(row=choice, column=1)
            b = ws.cell(row=choice, column=2)
            c = ws.cell(row=choice, column=3)
            d = ws.cell(row=choice, column=4)
            lastname.insert(0, a.value)
            firstname.insert(0, b.value)
            fathername.insert(0, c.value)
            age.insert(0, d.value)
            update_btn.place_forget()
            save_btn.place(width=100, height=20, x=725, y=170)

        else:
            mb.showerror('ОШИБКА', 'Выберите сотрудника из списка!')
    else:
        mb.showerror('ОШИБКА', 'Введите число в поле "номер сотрудника"')


def update():
    if lastname.get() and firstname.get() and fathername.get() and age.get():
        a = ws.cell(row=choice, column=1)
        b = ws.cell(row=choice, column=2)
        c = ws.cell(row=choice, column=3)
        d = ws.cell(row=choice, column=4)
        a.value = lastname.get()
        b.value = firstname.get()
        c.value = fathername.get()
        d.value = age.get()
        wb.save(filename)
        mb.showinfo('Информация', 'Изменения сохранены')
        lastname.delete(0, END)
        firstname.delete(0, END)
        fathername.delete(0, END)
        age.delete(0, END)
        num_employee.delete(0, END)
        view()
        save_btn.place_forget()
        update_btn.place(width=100, height=20, x=725, y=170)
    else:
        mb.showerror('ОШИБКА', 'Все поля должны быть заполнены!')


def delete():
    view()
    if num_employee.get():
        num = int(num_employee.get())
        if num > 0 and num <= rows_count:
            a = ws.cell(row=num, column=1)
            confirm = mb.askyesno('Внимание!', 'Вы точно хотите удалить '
                                  F'сотрудника {a.value}?\nДанное действие '
                                  'невозможно отменить.')
            if confirm:
                ws.delete_rows(num, 1)
                wb.save(filename)
                mb.showinfo('Информация', 'Сотрудник удален')
                num_employee.delete(0, END)
                view()
            else:
                num_employee.delete(0, END)
        else:
            mb.showerror('ОШИБКА', 'Выберите сотрудника из списка!')
    else:
        mb.showerror('ОШИБКА', 'Введите число в поле "номер сотрудника"')


def searcher(x):
    view()
    if x == 'lastname':
        last = lastname.get()
        showwarn_text = 'Сотрудника с такой фамилией нет!'
        col = 1
    if x == 'one letter':
        last = one_let.get()
        showwarn_text = f'Сотрудника с фамилией, начинающейся\n на {last} нет!'
        col = 1
    if x == 'age':
        last = age.get()
        showwarn_text = f'Сотрудника с возрастом - {last} нет!'
        col = 4
    new_text = ''
    count = 0
    for row in range(1, rows_count + 1):
        a = ws.cell(row=row, column=col)
        if (x == 'lastname' or x == 'age') and a.value == last:
            count += 1
            new_text += f'{count}.'
            for i in range(1, 5):
                b = ws.cell(row=row, column=i)
                if i != 4:
                    new_text += f'{b.value}  '
                else:
                    new_text += f'{b.value}\n'
        if x == 'one letter' and a.value[0] == last:
            count += 1
            new_text += f'{count}.'
            for i in range(1, 5):
                b = ws.cell(row=row, column=i)
                if i != 4:
                    new_text += f'{b.value}  '
                else:
                    new_text += f'{b.value}\n'
    if count > 0:
        text_staff.delete('1.0', END)
        text_staff.insert(END, new_text)
        search_btn.place_forget()
        view_btn.place(width=130, height=20, x=710, y=150)
        filename = 'search' + f'{str(random())[2:]}.txt'
        with open(filename, 'a+', encoding='utf-8') as file:
            file.write(new_text)

        mb.showinfo('Информация', 'Результаты поиска успешно сохранены \n'
                    f'в файле {filename}')

    else:
        mb.showwarning('Не найден', showwarn_text)


def search_lastname():
    if lastname.get():
        searcher('lastname')
    else:
        mb.showerror('ОШИБКА', 'Введите фамилию сотрудника в поле "Фамилия"')


def view_search():
    view()
    view_btn.place_forget()
    search_btn.place(width=100, height=20, x=725, y=150)
    lastname.delete(0, END)


def search_one_let():
    if len(one_let.get()) == 1:
        searcher('one letter')
    else:
        mb.showerror(
            'ОШИБКА', 'Введите первую букву фамилии сотрудника в поле "1 буква"'
            )


def search_age():
    if age.get():
        searcher('age')
    else:
        mb.showerror(
            'ОШИБКА', 'Введите искомый возраст сотрудника в поле "Возраст"')


win = Tk()
win.title('Сотрудники')
win.geometry('900x334+180+150')
win.resizable (width=False, height=False)
frame_staff = Frame(master=win, relief=RIDGE, borderwidth=5, bg='white')
frame_staff.grid(row=0, column=0)
text_staff = Text(master=frame_staff, width=80,
                  height=20, bg='white', fg='black')
text_staff.bind("<Button-1>", 'disable_mouse_event')
text_staff.grid(row=0, column=0, sticky='news')
scrollbar_staff = Scrollbar(frame_staff, orient='vertical',
                            command=text_staff.yview)
scrollbar_staff.grid(row=0, column=1, sticky='ns')
text_staff.configure(yscrollcommand=scrollbar_staff.set)

actions = [
    'Выбрать действие...', 'Добавить', 'Редактировать', 'Удалить',
    'Найти сотрудника', 'Найти по первой букве', 'Найти по возрасту']
actions_var = StringVar(value=actions[0])
add_label = Label(text='Добавить сотрудника:', font=("Times", 15), justify=LEFT,
                  fg='black')
update_label = Label(text='Редактировать сотрудника:', font=("Times", 12),
                     justify=LEFT, fg='black')
del_label = Label(text='Удалить сотрудника:', font=("Times", 15), justify=LEFT,
                  fg='black')
search_lastname_label = Label(text='Поиск по фамилии:', font=("Times", 15),
                              justify=LEFT, fg='black')
search_one_let_label = Label(text='Поиск по 1 букве:', font=("Times", 15),
                             justify=LEFT, fg='black')
search_age_label = Label(text='Поиск по возрасту:', font=("Times", 15),
                         justify=LEFT, fg='black')
one_let_label = Label(text='1 буква', font=("Times", 10), justify=LEFT,
                      fg='black')
one_let = ttk.Entry()
lastname_label = Label(text='Фамилия', font=("Times", 10), justify=LEFT,
                       fg='black')
lastname = ttk.Entry()
firstname_label = Label(text='Имя', font=(
    "Times", 10), justify=LEFT, fg='black')
firstname = ttk.Entry()
fathername_label = Label(text='Отчество', font=("Times", 10), justify=LEFT,
                         fg='black')
fathername = ttk.Entry()
age_label = Label(text='Возраст', font=("Times", 10), justify=LEFT, fg='black')
age = ttk.Entry()
num_employees_label = Label(text='Номер сотрудника', font=("Times", 10),
                            justify=LEFT, fg='black')
num_employee = ttk.Entry()
update_helper = Label(text='Введите номер сотрудника и нажмите \nредактировать.'
                      'Затем отредактируйте \nинформацию и нажмите сохранить.',
                      font=("Times", 10), justify=LEFT, fg='green')

add_btn = Button(text='Добавить', bg='lightgreen', activebackground='red',
                 command=add)
del_btn = Button(text='Удалить', bg='lightgreen', activebackground='red',
                 command=delete)
update_btn = Button(text='Редактировать', bg='lightgreen', activebackground=
                    'red', command=choice_update)
save_btn = Button(text='Сохранить', bg='lightgreen', activebackground='red',
                  command=update)
search_btn = Button(text='Найти', bg='lightgreen', activebackground='red',
                    command=search_lastname)
search_one_let_btn = Button(text='Найти', bg='lightgreen', activebackground=
                            'red',command=search_one_let)
search_age_btn = Button(text='Найти', bg='lightgreen', activebackground='red',
                        command=search_age)
view_btn = Button(text='Показать весь список', bg='lightgreen',
                  activebackground='red', command=view_search)

combobox = ttk.Combobox(textvariable=actions_var, values=actions)
open_btn = ttk.Button(text="Открыть файл", command=open_file)
open_btn.place(x=750, y=150)

win.mainloop()
